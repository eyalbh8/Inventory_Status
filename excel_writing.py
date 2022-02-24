from openpyxl import load_workbook


def save_changes(medicine_index, date, quantity, min_quantity):
    excel_workbook = load_workbook("medicine_inventory.xlsx")
    sheet = excel_workbook["mainSheet"]
    sheet["B" + str(medicine_index)] = date
    sheet["C" + str(medicine_index)] = quantity
    sheet["D" + str(medicine_index)] = min_quantity
    excel_workbook.save("medicine_inventory.xlsx")


def get_max_row():
    excel_workbook = load_workbook("medicine_inventory.xlsx")
    sheet = excel_workbook.active
    print(sheet.max_row)
    return sheet.max_row


def save_new_medicine(medicine_index, name, date, quantity, min_quantity):
    excel_workbook = load_workbook("medicine_inventory.xlsx")
    sheet = excel_workbook["mainSheet"]
    sheet["A" + str(medicine_index)] = name
    sheet["B" + str(medicine_index)] = date
    sheet["C" + str(medicine_index)] = quantity
    sheet["D" + str(medicine_index)] = min_quantity
    excel_workbook.save("medicine_inventory.xlsx")
