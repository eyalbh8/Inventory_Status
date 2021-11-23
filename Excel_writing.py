from openpyxl import load_workbook

def Save_Changes(Medicine_index, Date, Quantity, MIN_Quantity):
    Excel_Workbook = load_workbook("Medicine_inventory.xlsx")
    sheet = Excel_Workbook["mainSheet"]
    sheet["B" + str(Medicine_index)] = Date
    sheet["C" + str(Medicine_index)] = Quantity
    sheet["D" + str(Medicine_index)] = MIN_Quantity
    Excel_Workbook.save("Medicine_inventory.xlsx")

def Max_Row():
    Excel_Workbook = load_workbook("Medicine_inventory.xlsx")
    sheet = Excel_Workbook.active
    return sheet.max_row


def Save_New_Medicine(Medicine_index, Name, Date, Quantity, MIN_Quantity):
    Excel_Workbook = load_workbook("Medicine_inventory.xlsx")
    sheet = Excel_Workbook["mainSheet"]
    sheet["A" + str(Medicine_index)] = Name
    sheet["B" + str(Medicine_index)] = Date
    sheet["C" + str(Medicine_index)] = Quantity
    sheet["D" + str(Medicine_index)] = MIN_Quantity
    Excel_Workbook.save("Medicine_inventory.xlsx")
